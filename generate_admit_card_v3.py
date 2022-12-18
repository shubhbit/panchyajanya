from openpyxl import load_workbook
import pandas as pd
import os
import time
import string
import random
import copy

#################################### CONFIGS ####################################
PRINT_LEN = 65
ROLL_BASE = 19220000
FILL_CHAR = "*"
# change the file name xlsx which you want to read
file = "Panchyajanya_latest.xlsx"
internal_output_file = "Admit_Cards_Internal_Use"
external_output_file = "Admit_Cards_External_Use"
duplicate_records_file = "Admit_Cards_duplicate"
BANNER = " पाञ्चजन्य सामान्य ज्ञान प्रतियोगिता - २०२२ "
BOTTOM = " प्रवेश पत्र "
#################################### CONFIGS ####################################


def random_string(chars):
    N = chars
    res = ''.join(random.choices(string.ascii_uppercase +
                                 string.digits, k=N))
    return res


columns_we_need = {"Name": 2, "Roll No.": 11, "Registration No.": 1, "School": 9,
                   "Class": 6, "Gender": 7, "Father's Name": 3, "Interest": 12, "Date of Birth": 8,
                   "Aadhar No. / School ID": 10}


def get_no_cols_rows(file):
    df = pd.read_excel(file)
    return len(df.columns), len(df)


NO_OF_COLS, NO_OF_ROWS = get_no_cols_rows(file)

duplicate_records = []
def if_student_unique(final_data, row):
    is_unique = True
    for gender in final_data:
        for student in final_data[gender]:
            if student.lower().strip() == row[columns_we_need["Name"]].lower().strip():
                if str(final_data[gender][student]["Aadhar No. / School ID"]).lower().strip() == str(row[columns_we_need["Aadhar No. / School ID"]]).lower().strip():
                    dup_student = copy.deepcopy(final_data[gender][student])
                    dup_student["Name"] = student
                    duplicate_records.append(dup_student)
                    is_unique = False
                    break
                else:
                    new_name = "{0}_{1}".format(
                        student, random_string(len(student)))
                    row.pop(columns_we_need["Name"])
                    row.insert(columns_we_need["Name"], new_name)
        if not is_unique:
            break
    return row, is_unique


def read_data(sheet_obj):
    final_data = {"Female": {}, "Male": {}}
    duplicate = 0
    for row_i in range(2, NO_OF_ROWS+2):
        rows = []
        for column_i in range(1, NO_OF_COLS+1):
            cell_obj = sheet_obj.cell(row=row_i, column=column_i)
            rows.append(cell_obj.value)
        row, is_unique = if_student_unique(final_data, rows)
        if is_unique:
            rows = row
            final_data[rows[columns_we_need['Gender']]][rows[columns_we_need['Name']]] = {
                k: rows[columns_we_need[k]] for k in columns_we_need.keys() if k != "Name"}
        else:
            duplicate += 1
    print("##################### Summary #####################")
    print("total unique students: ", len(
        final_data['Male'])+len(final_data['Female']))
    print("duplicates removed: ", duplicate)
    return final_data


def group_by_interest(final_data):
    data_by_interest = {"Female": {}, "Male": {}}
    for gender in final_data:
        for student in sorted(final_data[gender]):
            interest = final_data[gender][student]['Interest']
            if interest not in data_by_interest[gender]:
                data_by_interest[gender][interest] = [
                    {student: final_data[gender][student]}]
            else:
                data_by_interest[gender][interest].append(
                    {student: final_data[gender][student]})
    return data_by_interest


def assign_roll_number(student_by_interest):
    global ROLL_BASE
    for gender in student_by_interest:
        for students in student_by_interest[gender].values():
            for student in students:
                ROLL_BASE += 1
                for key in student:
                    student[key]["Roll No."] = ROLL_BASE
    return student_by_interest


def group_by_school(student_by_with_role):
    students_by_school = {}
    for gender in student_by_with_role:
        for students in student_by_with_role[gender].values():
            for student in students:
                for key in student:
                    student[key]["Name"] = key
                    if student[key]["School"] not in students_by_school:
                        students_by_school[student[key]
                                           ["School"]] = [student[key]]
                    else:
                        students_by_school[student[key]
                                           ["School"]].append(student[key])
    return students_by_school


def cast_float_to_int(key, value):
    if key in ["Registration No.", "Aadhar No. / School ID", "Class"]:
        if type(value) == float:
            return int(value)
    else:
        return value


def generate_print_data(data_by_school, template, external=False):
    output_file = internal_output_file
    if external:
        output_file = external_output_file
    if os.path.exists(output_file):
        os.remove(output_file)
    for school in data_by_school:
        school_data = data_by_school[school]
        for student in sorted(school_data, key=lambda x: x['Class']):
            data_list = []
            data_list.append(BANNER.center(PRINT_LEN+8, "*"))
            name_to_print = student["Name"].split("_")[0]
            data_list.append("Name: {}".format(
                name_to_print).center(PRINT_LEN, " "))
            del student["Name"]
            if external:
                output_file = external_output_file
                del student["Aadhar No. / School ID"]
            sorted_keys = list(columns_we_need.keys())
            for indx, k in enumerate(sorted_keys):
                if k != "Aadhar No. / School ID" and k != "Name":
                    if indx % 2 != 0:
                        v = student[k]
                        if k == "Date of Birth":
                            v = str(v).split(" ")[0]
                        v = cast_float_to_int(k, v)
                        fi = "{}: {}".format(k, v)
                        new_key = sorted_keys[indx+1]
                        new_value = student[new_key]
                        if new_key == "Date of Birth":
                            new_value = str(new_value).split(" ")[0]
                        new_value = cast_float_to_int(new_key, new_value)
                        se = " {}: {}".format(
                            new_key, new_value)
                        space = PRINT_LEN - (len(fi)+len(se))
                        temp_d = "{}{}{}".format(fi.ljust(0), " "*space, se)
                        data_list.append(temp_d)
            if not external:
                identity = student["Aadhar No. / School ID"]
                data_list.append("Aadhar No. / School ID: {}".format(
                    identity).center(PRINT_LEN, " "))
            data_list.append(BOTTOM.center(PRINT_LEN, "*"))
            with open(output_file, "a") as f:
                f.write(template.format(*data_list))

def write_duplicate_record():
    global ROLL_BASE
    global duplicate_records
    if os.path.exists(duplicate_records_file):
        os.remove(duplicate_records_file)
    with open(duplicate_records_file, "a") as f:
        for student in duplicate_records:
            ROLL_BASE += 1
            student["Roll No."] = ROLL_BASE
            f.write("{}\n".format(BANNER.center(PRINT_LEN+8, "*")))
            for key in columns_we_need.keys():
                value = cast_float_to_int(key, student[key])
                if key == "Date of Birth":
                    value = str(value).split(" ")[0]
                st = "{}: {}".format(key, value)
                f.write(st.center(PRINT_LEN, " "))
                f.write("\n")
            f.write("{}".format(BOTTOM.center(PRINT_LEN, "*")))
            f.write("\n\n")



def print_new(data_by_school):
    st_internal = """
{0}\n
{1}\n
{2}\n
{3}
{4}
{5}
{6}\n
{7}
"""
    st_external = """
{0}\n
{1}\n
{2}\n
{3}\n
{4}\n
{5}
{6}
"""
    generate_print_data(copy.deepcopy(data_by_school), st_external, True)
    generate_print_data(copy.deepcopy(data_by_school), st_internal)
    write_duplicate_record()


def main():
    wb_obj = load_workbook(file)
    sheet_obj = wb_obj.active
    start_time = time.time()
    final_data = read_data(sheet_obj)
    student_by_interest = group_by_interest(final_data)
    students_with_roll = assign_roll_number(student_by_interest)
    students_by_school = group_by_school(students_with_roll)
    end_time = time.time()
    print_new(students_by_school)
    print("time took: {} seconds".format(end_time-start_time))
    print("##################### Summary #####################")


if __name__ == "__main__":
    main()
